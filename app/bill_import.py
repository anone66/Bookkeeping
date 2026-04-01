"""支付宝 CSV / 微信 XLSX 账单解析（与仓库 `账单/` 样例格式对齐）。"""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass
from datetime import datetime, timedelta
from typing import Any

ALI_PAY_HEADER_FIRST = "交易时间"
WECHAT_HEADER_FIRST = "交易时间"

# 支付宝：仅「交易成功」；微信：样例中出现的成功类状态
ALIPAY_OK_STATUS = frozenset({"交易成功"})
WECHAT_OK_STATUS = frozenset(
    {
        "支付成功",
        "已转账",
        "交易成功",
        "充值完成",
        "对方已收钱",
        "已存入零钱",
    }
)


def _strip_id(s: str) -> str:
    return re.sub(r"\s+", "", (s or "").strip())


def _parse_alipay_amount(s: str) -> float | None:
    t = (s or "").strip().replace(",", "")
    if not t:
        return None
    try:
        return float(t)
    except ValueError:
        return None


def _parse_wechat_amount(s: str) -> float | None:
    return _parse_alipay_amount(s)


def _alipay_datetime_to_date(s: str) -> str | None:
    t = (s or "").strip()
    if not t:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S"):
        try:
            return datetime.strptime(t[:19], fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    try:
        return datetime.strptime(t[:10], "%Y-%m-%d").strftime("%Y-%m-%d")
    except ValueError:
        return None


def _excel_serial_to_date(val: float) -> str:
    """Excel 序列日 → YYYY-MM-DD（与 openpyxl/Excel Windows 纪元一致）。"""
    base = datetime(1899, 12, 30)
    dt = base + timedelta(days=float(val))
    return dt.strftime("%Y-%m-%d")


def _wechat_time_to_date(val: Any) -> str | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, (int, float)):
        return _excel_serial_to_date(float(val))
    s = str(val).strip()
    return _alipay_datetime_to_date(s)


@dataclass
class BillImportRow:
    type: str  # expense | income
    amount_cents: int
    transacted_on: str
    external_id: str
    bill_category: str
    bill_counterparty: str
    bill_product: str
    bill_payment_method: str
    bill_merchant_no: str
    bill_export_note: str


@dataclass
class BillParseResult:
    rows: list[BillImportRow]
    skipped_neutral: int
    skipped_bad_status: int
    skipped_zero_amount: int
    skipped_no_date: int
    skipped_no_external_id: int


def _decode_csv(raw: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def parse_alipay_csv(raw: bytes) -> tuple[BillParseResult | None, str | None]:
    """解析支付宝交易明细 CSV。失败时 (None, error_message)。"""
    text = _decode_csv(raw)
    reader = csv.reader(io.StringIO(text))
    rows_list = list(reader)
    header_idx = None
    for i, row in enumerate(rows_list):
        if row and row[0].strip("\ufeff ").strip() == ALI_PAY_HEADER_FIRST:
            header_idx = i
            break
    if header_idx is None:
        return None, "未找到支付宝明细表头（首列应为「交易时间」）"

    header = [h.strip() for h in rows_list[header_idx]]

    def col(name: str) -> int | None:
        try:
            return header.index(name)
        except ValueError:
            return None

    idx = {
        "time": col("交易时间"),
        "cat": col("交易分类"),
        "who": col("交易对方"),
        "product": col("商品说明"),
        "flow": col("收/支"),
        "amt": col("金额"),
        "pay": col("收/付款方式"),
        "status": col("交易状态"),
        "tid": col("交易订单号"),
        "mid": col("商家订单号"),
        "note": col("备注"),
    }
    if idx["time"] is None or idx["flow"] is None or idx["amt"] is None:
        return None, "支付宝 CSV 缺少必要列（交易时间、收/支、金额）"

    def alipay_cell(row: list[str], key: str) -> str:
        j = idx.get(key)
        if j is None or j >= len(row):
            return ""
        return (row[j] or "").strip()

    out: list[BillImportRow] = []
    skipped_neutral = 0
    skipped_bad_status = 0
    skipped_zero_amount = 0
    skipped_no_date = 0
    skipped_no_external_id = 0

    for row in rows_list[header_idx + 1 :]:
        if not row or all(not (c or "").strip() for c in row):
            continue

        flow = alipay_cell(row, "flow")
        if flow == "不计收支":
            skipped_neutral += 1
            continue
        if flow not in ("支出", "收入"):
            continue

        status = alipay_cell(row, "status")
        if status not in ALIPAY_OK_STATUS:
            skipped_bad_status += 1
            continue

        amt = _parse_alipay_amount(alipay_cell(row, "amt"))
        if amt is None or amt <= 0:
            skipped_zero_amount += 1
            continue

        transacted_on = _alipay_datetime_to_date(alipay_cell(row, "time"))
        if not transacted_on:
            skipped_no_date += 1
            continue

        ext = _strip_id(alipay_cell(row, "tid"))
        if not ext:
            skipped_no_external_id += 1
            continue

        amount_cents = int(round(amt * 100))
        if amount_cents <= 0:
            skipped_zero_amount += 1
            continue

        typ = "expense" if flow == "支出" else "income"
        mid_s = alipay_cell(row, "mid")
        out.append(
            BillImportRow(
                type=typ,
                amount_cents=amount_cents,
                transacted_on=transacted_on,
                external_id=ext,
                bill_category=alipay_cell(row, "cat"),
                bill_counterparty=alipay_cell(row, "who"),
                bill_product=alipay_cell(row, "product"),
                bill_payment_method=alipay_cell(row, "pay"),
                bill_merchant_no=_strip_id(mid_s) or mid_s.strip(),
                bill_export_note=alipay_cell(row, "note"),
            )
        )

    return (
        BillParseResult(
            rows=out,
            skipped_neutral=skipped_neutral,
            skipped_bad_status=skipped_bad_status,
            skipped_zero_amount=skipped_zero_amount,
            skipped_no_date=skipped_no_date,
            skipped_no_external_id=skipped_no_external_id,
        ),
        None,
    )


def parse_wechat_xlsx(raw: bytes) -> tuple[BillParseResult | None, str | None]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None, "服务器未安装 openpyxl，无法解析微信账单"

    bio = io.BytesIO(raw)
    try:
        wb = load_workbook(bio, read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        return None, f"无法读取 Excel 文件：{e}"

    header_idx = None
    header: list[str] = []
    rows_iter = list(ws.iter_rows(values_only=True))
    wb.close()

    for i, row in enumerate(rows_iter):
        if not row:
            continue
        first = row[0]
        if first is not None and str(first).strip() == WECHAT_HEADER_FIRST:
            header_idx = i
            header = [str(c).strip() if c is not None else "" for c in row]
            break

    if header_idx is None:
        return None, "未找到微信明细表头（首列应为「交易时间」）"

    def col(name: str) -> int | None:
        try:
            return header.index(name)
        except ValueError:
            return None

    idx = {
        "time": col("交易时间"),
        "cat": col("交易类型"),
        "who": col("交易对方"),
        "product": col("商品"),
        "flow": col("收/支"),
        "amt": col("金额(元)"),
        "pay": col("支付方式"),
        "status": col("当前状态"),
        "tid": col("交易单号"),
        "mid": col("商户单号"),
        "note": col("备注"),
    }
    if idx["time"] is None or idx["flow"] is None or idx["amt"] is None:
        return None, "微信 XLSX 缺少必要列（交易时间、收/支、金额(元)）"

    def wechat_cell(row: tuple[Any, ...], key: str) -> str:
        j = idx.get(key)
        if j is None or j >= len(row):
            return ""
        v = row[j]
        if v is None:
            return ""
        return str(v).strip()

    out: list[BillImportRow] = []
    skipped_neutral = 0
    skipped_bad_status = 0
    skipped_zero_amount = 0
    skipped_no_date = 0
    skipped_no_external_id = 0

    for row in rows_iter[header_idx + 1 :]:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue

        flow = wechat_cell(row, "flow")
        if flow == "中性交易":
            skipped_neutral += 1
            continue
        if flow not in ("支出", "收入"):
            continue

        status = wechat_cell(row, "status")
        if status not in WECHAT_OK_STATUS:
            skipped_bad_status += 1
            continue

        amt = _parse_wechat_amount(wechat_cell(row, "amt"))
        if amt is None or amt <= 0:
            skipped_zero_amount += 1
            continue

        tval = row[idx["time"]] if idx["time"] < len(row) else None
        transacted_on = _wechat_time_to_date(tval)
        if not transacted_on:
            skipped_no_date += 1
            continue

        ext = _strip_id(wechat_cell(row, "tid"))
        if not ext:
            skipped_no_external_id += 1
            continue

        amount_cents = int(round(amt * 100))
        if amount_cents <= 0:
            skipped_zero_amount += 1
            continue

        typ = "expense" if flow == "支出" else "income"
        mid_raw = wechat_cell(row, "mid")
        out.append(
            BillImportRow(
                type=typ,
                amount_cents=amount_cents,
                transacted_on=transacted_on,
                external_id=ext,
                bill_category=wechat_cell(row, "cat"),
                bill_counterparty=wechat_cell(row, "who"),
                bill_product=wechat_cell(row, "product"),
                bill_payment_method=wechat_cell(row, "pay"),
                bill_merchant_no=_strip_id(mid_raw) or mid_raw,
                bill_export_note=wechat_cell(row, "note"),
            )
        )

    return (
        BillParseResult(
            rows=out,
            skipped_neutral=skipped_neutral,
            skipped_bad_status=skipped_bad_status,
            skipped_zero_amount=skipped_zero_amount,
            skipped_no_date=skipped_no_date,
            skipped_no_external_id=skipped_no_external_id,
        ),
        None,
    )
